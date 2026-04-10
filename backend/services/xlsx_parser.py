# -*- coding: utf-8 -*-
"""
xlsx_parser.py — 微信导出 xlsx 文件解析器

支持格式：
- 列：time / speaker / content（三列顺序可能不同）
- speaker：对方名字 或 用户手机号
- content：文本内容，可能包含 XML（图片/表情等）

输出：
- contact_name: str
- messages: list[dict]  [{speaker, content, time}]
"""

import hashlib
import openpyxl
from pathlib import Path
from typing import BinaryIO, NamedTuple


class ParseResult(NamedTuple):
    contact_name: str
    messages: list[dict]
    message_count: int
    file_hash: str


def compute_hash(file_content: bytes) -> str:
    return hashlib.sha256(file_content).hexdigest()


def parse(file_path_or_stream, filename: str = "") -> ParseResult:
    """
    解析微信导出的 xlsx 文件。

    Args:
        file_path_or_stream: 文件路径（str/Path）或文件上传的二进制流
        filename: 原始文件名，用于错误提示

    Returns:
        ParseResult(contact_name, messages, message_count, file_hash)

    Raises:
        ValueError: 文件格式不正确或无法解析
    """
    if isinstance(file_path_or_stream, (str, Path)):
        wb = openpyxl.load_workbook(str(file_path_or_stream), read_only=True, data_only=True)
        file_content = open(str(file_path_or_stream), "rb").read()
    else:
        # 二进制流
        file_content = file_path_or_stream.read()
        wb = openpyxl.load_workbook(file_content, read_only=True, data_only=True)

    file_hash = compute_hash(file_content)

    ws = wb.active
    headers = None
    col_time = col_speaker = col_content = -1

    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        raise ValueError(f"文件为空{f' ({filename})' if filename else ''}")

    # 找表头行
    header_row_idx = None
    for i, row in enumerate(rows[:10]):
        if row and any(cell is not None for cell in row):
            # 常见表头关键词
            row_lower = [str(c).lower() if c is not None else "" for c in row]
            if any("time" in c or "时间" in c for c in row_lower):
                if any("speaker" in c or "发送者" in c or "内容" in c or "content" in c for c in row_lower):
                    header_row_idx = i
                    headers = row_lower
                    break

    if header_row_idx is None:
        # 尝试列顺序推断（默认：time, speaker, content）
        headers = ["time", "speaker", "content"]
        header_row_idx = 0

    # 建立列索引映射
    for i, h in enumerate(headers):
        h = h.strip().lower()
        if "time" in h or "时间" in h or "日期" in h:
            col_time = i
        elif "speaker" in h or "发送者" in h or "nick" in h or "昵称" in h:
            col_speaker = i
        elif "content" in h or "内容" in h or "消息" in h:
            col_content = i

    # 如果没找到精确匹配，尝试列位置
    if col_time == -1 or col_speaker == -1 or col_content == -1:
        if len(rows[header_row_idx]) >= 3:
            # 默认：第1列=time, 第2列=speaker, 第3列=content
            col_time = 0 if col_time == -1 else col_time
            col_speaker = 1 if col_speaker == -1 else col_speaker
            col_content = 2 if col_content == -1 else col_content

    messages = []
    contact_name = ""
    other_speakers = set()

    for row in rows[header_row_idx + 1:]:
        if not row or all(cell is None for cell in row):
            continue

        time_val = row[col_time] if col_time < len(row) else None
        speaker = row[col_speaker] if col_speaker < len(row) else None
        content = row[col_content] if col_content < len(row) else None

        if not content or not str(content).strip():
            continue

        content_str = str(content).strip()

        # 跳过系统消息
        skip_keywords = [
            "以上是聊天记录",
            "%%emoji",
            "图片",
            "[图片]",
            "语音",
            "[语音]",
            "表情",
            "[表情包]",
            "file",
            "[文件]",
        ]
        if any(kw in content_str for kw in skip_keywords):
            continue
        if len(content_str) < 2:
            continue

        speaker_str = str(speaker).strip() if speaker else "未知"

        # 收集非"我"说话的人名
        if speaker_str not in ("我", "me", "Me", "ME", "自己"):
            if speaker_str not in ("未知", ""):
                other_speakers.add(speaker_str)

        messages.append({
            "speaker": speaker_str,
            "content": content_str,
            "time": str(time_val) if time_val else "",
        })

    if not messages:
        raise ValueError(f"未找到有效聊天记录{f' ({filename})' if filename else ''}")

    # 对方名字：取出现最多的非"我"说话人
    if other_speakers:
        speaker_counts = {}
        for m in messages:
            s = m["speaker"]
            if s not in ("我", "me", "Me", "ME", "自己"):
                speaker_counts[s] = speaker_counts.get(s, 0) + 1
        if speaker_counts:
            contact_name = max(speaker_counts, key=speaker_counts.get)
        else:
            contact_name = list(other_speakers)[0] if other_speakers else "未知联系人"
    else:
        contact_name = "未知联系人"

    return ParseResult(
        contact_name=contact_name,
        messages=messages,
        message_count=len(messages),
        file_hash=file_hash,
    )


def sample_messages(messages: list[dict], max_count: int = 500) -> str:
    """
    取最近 max_count 条消息，格式化为 LLM 输入。
    保留时间顺序（ oldest → newest）。
    """
    # 消息已经是按时间顺序的，取中间段（最能体现风格）
    total = len(messages)
    if total <= max_count:
        sampled = messages
    else:
        # 均匀采样：头尾各留一部分，中间均匀采样
        head = messages[:50]
        tail = messages[-50:]
        step = total / max_count
        body = [messages[int(i * step)] for i in range(50, total - 50) if int(i * step) < total - 50]
        sampled = head + body + tail

    lines = []
    for m in sampled:
        speaker_label = "对方" if m["speaker"] not in ("我", "me", "Me", "ME", "自己") else "我"
        lines.append(f"{speaker_label}: {m['content']}")
    return "\n".join(lines)
