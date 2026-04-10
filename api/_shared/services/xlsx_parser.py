# -*- coding: utf-8 -*-
"""
xlsx_parser.py — 微信导出 xlsx 文件解析器（Serverless 版本）
"""
import hashlib
import openpyxl
from pathlib import Path
from typing import BinaryIO, NamedTuple

class ParseResult(NamedTuple):
    contact_name: str
    messages: list
    message_count: int
    file_hash: str

def compute_hash(file_content: bytes) -> str:
    return hashlib.sha256(file_content).hexdigest()

def parse(file_path_or_stream, filename: str = "") -> ParseResult:
    if isinstance(file_path_or_stream, (str, Path)):
        wb = openpyxl.load_workbook(str(file_path_or_stream), read_only=True, data_only=True)
        file_content = open(str(file_path_or_stream), "rb").read()
    else:
        file_content = file_path_or_stream.read()
        wb = openpyxl.load_workbook(file_content, read_only=True, data_only=True)

    file_hash = compute_hash(file_content)
    ws = wb.active
    headers = None
    col_time = col_speaker = col_content = -1

    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        raise ValueError(f"文件为空{f' ({filename})' if filename else ''}")

    header_row_idx = None
    for i, row in enumerate(rows[:10]):
        if row and any(cell is not None for cell in row):
            row_lower = [str(c).lower() if c is not None else "" for c in row]
            if any("time" in c or "时间" in c for c in row_lower):
                if any("speaker" in c or "发送者" in c or "内容" in c or "content" in c for c in row_lower):
                    header_row_idx = i
                    headers = row_lower
                    break

    if header_row_idx is None:
        headers = ["time", "speaker", "content"]
        header_row_idx = 0

    for i, h in enumerate(headers):
        h = h.strip().lower()
        if "time" in h or "时间" in h or "日期" in h:
            col_time = i
        elif "speaker" in h or "发送者" in h or "nick" in h or "昵称" in h:
            col_speaker = i
        elif "content" in h or "内容" in h or "消息" in h:
            col_content = i

    if col_time == -1 or col_speaker == -1 or col_content == -1:
        if len(rows[header_row_idx]) >= 3:
            col_time = 0 if col_time == -1 else col_time
            col_speaker = 1 if col_speaker == -1 else col_speaker
            col_content = 2 if col_content == -1 else col_content

    messages = []
    other_speakers = set()

    for row in rows[header_row_idx + 1:]:
        if not row or all(cell is None for cell in row):
            continue
        time_val = row[col_time] if col_time < len(row) else None
        speaker = row[col_speaker] if col_speaker < len(row) else None
        content = row[col_content] if col_content < len(row) else None
        if not content or not str(content).strip():
            continue
        content_str = str(content).strip()
        skip_keywords = ["以上是聊天记录", "%%emoji", "图片", "[图片]", "语音", "[语音]", "表情", "[表情包]", "file", "[文件]"]
        if any(kw in content_str for kw in skip_keywords):
            continue
        if len(content_str) < 2:
            continue
        speaker_str = str(speaker).strip() if speaker else "未知"
        if speaker_str not in ("我", "me", "Me", "ME", "自己"):
            if speaker_str not in ("未知", ""):
                other_speakers.add(speaker_str)
        messages.append({"speaker": speaker_str, "content": content_str, "time": str(time_val) if time_val else ""})

    if not messages:
        raise ValueError(f"未找到有效聊天记录{f' ({filename})' if filename else ''}")

    contact_name = "未知联系人"
    if other_speakers:
        speaker_counts = {}
        for m in messages:
            s = m["speaker"]
            if s not in ("我", "me", "Me", "ME", "自己"):
                speaker_counts[s] = speaker_counts.get(s, 0) + 1
        if speaker_counts:
            contact_name = max(speaker_counts, key=speaker_counts.get)
        else:
            contact_name = list(other_speakers)[0] if other_speakers else "未知联系人"

    return ParseResult(contact_name=contact_name, messages=messages, message_count=len(messages), file_hash=file_hash)

def sample_messages(messages: list, max_count: int = 500) -> str:
    total = len(messages)
    if total <= max_count:
        sampled = messages
    else:
        head = messages[:50]
        tail = messages[-50:]
        step = total / max_count
        body = [messages[int(i * step)] for i in range(50, total - 50) if int(i * step) < total - 50]
        sampled = head + body + tail
    lines = []
    for m in sampled:
        speaker_label = "对方" if m["speaker"] not in ("我", "me", "Me", "ME", "自己") else "我"
        lines.append(f"{speaker_label}: {m['content']}")
    return "\n".join(lines)
